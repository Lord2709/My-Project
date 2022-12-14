# Libraries...
import os
os.system('cls')
import shutil
import pandas as pd
import streamlit as st
from PIL import Image
import openpyxl
import math
import numpy as np
from openpyxl.styles import Alignment,Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import datetime
import re
import time
import matplotlib.pyplot as plt
import plotly.graph_objects as go
from zipfile import ZipFile
from os.path import basename


from datetime import datetime
start_time = datetime.now()


# write_in_xlsx function is used to append values column wise 
try:
    def write_in_xlsx(range_of_rows,column_number,ls,ws):
        for i in range(1,range_of_rows+1):
            cellref=ws.cell(row=i, column=column_number)
            if i == 1:
                cellref.value = ls[0]
            else:
                cellref.value=ls[i-1]
except:
	print("Error in calling write_in_xlsx function")



# fontsize function is used to change the font size 
try:
    def fontsize(size,ws):
        mr = ws.max_row
        mc = ws.max_column
        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                c = ws.cell(row = i, column = j)
                c.font = Font(size = size)
except:
	print("Error in calling fontsize function")



# calculate_rank function is used to calculate rank part of the table in tut05 part...
try:
    def calculate_rank(vector):
        a={}
        rank=1
        for num in sorted(vector, reverse=True):
            if num not in a:
                a[num]=rank
                rank=rank+1
        return[a[i] for i in vector]
except:
	print("Error in calling calculate_rank function")



# set_border function is used to apply border to the tables...
try:
    def set_border(ws, cell_range):
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
except:
	print("Error in calling set_border function")



# bgcolor function is used to apply background color to the table by conditional formatting...
try:
    def bgcolor(range, fill_color,ws,temp):
        fill = PatternFill(start_color=fill_color,end_color=fill_color,fill_type='solid')


        if temp == 0:
            for row in ws[range]:
                count = 0
                for cell in row:
                    if cell.value == 1 and count == 0:
                        cell.fill = fill
                        count = 1
            # ws.conditional_formatting.add(range, CellIsRule(operator='equal', formula=[1], fill=fill))
        elif temp == 1 or temp == 3:
            for row in ws[range]:
                for cell in row:
                    cell.fill = fill
        elif temp == 2:
            for row in ws[range]:
                row[0].fill = fill
except:
	print("Error in calling bgcolor function")



# max_color is used to color max value in the row in transition count table by conditional formatting...
try:
    def max_color(range1,ws,color):
        if color != 0:
            fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
        else:
            fill = PatternFill(start_color='FFFF00',end_color='FFFF00',fill_type='solid')

        max_list = []

        for row in ws[range1]:
            list_row_val = []
            for cell in row:
                list_row_val.append(cell.value)
            max_list.append(max(list_row_val))
      
        for row in ws[range1]:
            for cell in row:
                if cell.value == max_list[0]:
                    cell.fill = fill
                    if len(max_list) != 0:
                        del max_list[0]
                    break
                else:
                    continue
except:
	print("Error in calling max_color function")                



try:
    def zip_file(current_path, iter):
        if iter:
            os.chdir(current_path.rsplit("/input")[0])
        else:
            os.chdir(current_path)

        # create a ZipFile object
        dirName = './output3'
        with ZipFile('Octant.zip', 'w') as zipObj:
        # Iterate over all the files in directory
            for folderName, subfolders, filenames in os.walk(dirName):
                for filename in filenames:
                    #create complete filepath of file in directory
                    filePath = os.path.join(folderName, filename)
                    # Add file to zip
                    zipObj.write(filePath, basename(filePath))


        st.success("Download zip file from below")
        col1, col2, col3 = st.columns(3)
        with open("Octant.zip", "rb") as fp:
            btn = col2.download_button(
                label="Download ZIP",
                data=fp,
                file_name="Octant.zip",
                mime="application/zip"
            )

        shutil.rmtree('./output3')
        time.sleep(60)
except:
	print("Error in calling zip_file function")

    


# global output_file_name
try:
    def octant_analysis(iter,path,mod):
        # 2 cases are present, one for single file (iter = 0) and another is bulk conversion (iter =1)
        
        if iter==1:
            current_dir = path.rsplit("/",1)[0]
            os.chdir(current_dir)

            # if output folder exists with unwanted files, then deleting the files and then the dir...
            if os.path.isdir("output3"):
                list_of_files = os.listdir('output3')
                for files in list_of_files:
                    os.remove(current_dir+"/output3/" + f"{files}")
                os.rmdir("output3")


            os.mkdir(current_dir + "/output3/")
            output_dir = current_dir + "/output3/"

            # Separately stored the input file names in files variable
            files = os.listdir(path)
        else:
            if os.getcwd().rsplit("\\")[-1] == "output3":
                current_dir = os.getcwd()[:os.getcwd().index('\output3')].replace("\\","/")

                output_dir = current_dir + "/output3/"

                list_of_files = os.listdir(current_dir + "/output3")
                for files in list_of_files:
                    os.remove(output_dir + f"{files}")

            else:
                current_dir = os.getcwd().replace("\\","/")
                output_dir = current_dir + "/output3/"
                if os.path.exists(current_dir + "/output3"):
                    pass
                else:
                    os.mkdir(output_dir)

            files = [path[0]]
            filename = path[1]

        octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}


        # "for" loop to iterate over all input files present in the folder or over the uploaded file.
        for f in files:
            
            if iter==1:
                file_path = path + '/' + f

            #====================================
            # For Single File Operation
            if iter != 1:
                wb_output = openpyxl.Workbook()
                ws_output = wb_output.active

                A = ['','T'] + list(f['T'])
                B = ['','U'] + list(f['U'])
                C = ['','V'] + list(f['V'])
                D = ['','W'] + list(f['W'])

                write_in_xlsx(len(A), 1, A, ws_output)
                write_in_xlsx(len(A), 2, B, ws_output)
                write_in_xlsx(len(A), 3, C, ws_output)
                write_in_xlsx(len(A), 4, D, ws_output)

            #====================================
            # For Bulk Conversion
            else:
                wb = openpyxl.load_workbook(file_path)
                ws = wb.active
                rows = ws.iter_rows(min_row = ws.min_row, max_row=ws.max_row, min_col = ws.min_column, max_col = ws.max_column)

                # Setting first 4 columns copied from input file to output file
                wb_output = openpyxl.Workbook()
                ws_output = wb_output.active
                mr = ws.max_row
                mc = ws.max_column
                for i in range (1, mr + 1):
                    for j in range (1, mc + 1):
                        c = ws.cell(row = i, column = j)
                        ws_output.cell(row = i+1, column = j).value = c.value
                # ===============================================================

            os.chdir(output_dir)

            output = [[],[],[],[]]
            l_col = ['A','B','C','D']
            for i in range(4):
                col = ws_output[l_col[i]][2:]
                for cell in col:
                    output[i].append(cell.value)
    


            # Calculating U,V,W average and storing in the output file 
            ws_output['E1'].value = " "
            ws_output['E2'].value = "U Avg"
            ws_output['E3'].value = np.round_(np.mean(output[1]), decimals = 3)
            ws_output['F1'].value = " "
            ws_output['F2'].value = "V Avg"
            ws_output['F3'].value = np.round_(np.mean(output[2]), decimals = 3)
            ws_output['G1'].value = " "
            ws_output['G2'].value = "W Avg"
            ws_output['G3'].value = np.round_(np.mean(output[3]), decimals = 3)
            # =====================================================


            # Updating H, I, J column of output file
            H = output[1] - ws_output['E3'].value
            I = output[2] - ws_output['F3'].value
            J = output[3] - ws_output['G3'].value
            H_col = [" ","U'=U-U Avg"]
            I_col = [" ","V'=V-V Avg"]
            J_col = [" ","W'=W-W Avg"]
            for i in range(len(H)):
                H_col.append(np.round_((H[i]), decimals = 3))
                I_col.append(np.round_((I[i]), decimals = 3))
                J_col.append(np.round_((J[i]), decimals = 3))
        
            len_H = len(H_col)
            write_in_xlsx(len_H, 8, H_col, ws_output)
            write_in_xlsx(len_H, 9, I_col, ws_output)
            write_in_xlsx(len_H, 10, J_col, ws_output)
            # =============================================


            # Updating Octant column of output file
            Oct = [" ","Octant"]
            for i in range(len_H):
                if i==0 or i==1: 
                    continue
                else:
                    if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(1)
                    if((H_col[i]>=0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-1)
                    if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]>=0)): Oct.append(2)
                    if((H_col[i]<0) & (I_col[i]>=0) & (J_col[i]<0)): Oct.append(-2)
                    if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(3)
                    if((H_col[i]<0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-3)
                    if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]>=0)): Oct.append(4)
                    if((H_col[i]>=0) & (I_col[i]<0) & (J_col[i]<0)): Oct.append(-4)
            write_in_xlsx(len_H, 11, Oct, ws_output)
            # ==========================================
            

            # ========================================== Tut05 Part-Start ===============================================

            M_col = ['','','',f'Mod {mod}']
            write_in_xlsx(len(M_col), 13, M_col, ws_output)
            
            N_col = ["Overall Octant Count","", "Octant ID", "Overall Count"]

            n = len(Oct[2:])
            if n%mod == 0:
                n_ranges = n//mod
            else:
                n_ranges = math.ceil(n/mod)


            t = 0
            u = t + mod
            for j in range(n_ranges):
                if u <= n:
                    if t==0:
                        N_col.append(f"0000 - {u-1}")
                    else:
                        N_col.append(f"{t} - {u-1}")
                else:
                    N_col.append(f"{t} - {n}")
                t = u
                u += mod
            write_in_xlsx(len(N_col), 14, N_col, ws_output)

            # 8 new list are created to store the count value of {+1,-2,+2,-2,+3,-3,+4,-4}
            overall_count = []
            e = 0
            st = ["+1","-1","+2","-2","+3","-3","+4","-4"]
            for i in range(8):
                empty_list = ["",""]
                empty_list.append(st[e])
                overall_count.append(empty_list)
                e += 1

            df2 = Oct[2:]
            int_l = [1,-1,2,-2,3,-3,4,-4]
            j = 0
            for i in range(8):
                overall_count[i].append(df2.count(int_l[j]))
                j += 1
            # n --> given in tut01.pdf that max value will never exceed 30000.
            # Below while function is used to count the {+1,-2,+2,-2,+3,-3,+4,-4} values for different mod's

            k = 0
            m = mod
            n_r = n_ranges
            while n_r > 0:   
                n_r -= 1
                d_f = df2[k:m]
                j = 0
                for i in range(8):
                    overall_count[i].append(d_f.count(int_l[j]))
                    j += 1
                k = m 
                m = m + mod


            col = 15
            for i in range(8):
                write_in_xlsx(len(overall_count[i]), col, overall_count[i], ws_output)
                col += 1


            overall_rank = [["","","Rank Octant 1"],["","","Rank Octant -1"],["","","Rank Octant 2"],["","","Rank Octant -2"],["","","Rank Octant 3"],["","","Rank Octant -3"],["","","Rank Octant 4"],["","","Rank Octant -4"]]
            AC_col = ["","Octant ID"] + int_l
            AD_col = ["","Octant Name","Internal outward interaction","External outward interaction",
                    "External Ejection","Internal Ejection","External inward interaction",
                    "Internal inward interaction","Internal sweep","External sweep"]

            row_list = []
            imp_element_pos = [3] + [i for i in range(4,n_ranges+4)]
            for i in imp_element_pos:
                l = []
                for j in range(8):
                    l.append(overall_count[j][i])
                row_list.append(l)

        
            rank_list = []
            for i in range(len(row_list)):
                rank_list.append(calculate_rank(row_list[i]))

    
            for i in range(len(overall_rank)):
                for j in rank_list:
                    overall_rank[i].append(j[i])

            overall_rank[6] += AC_col
            overall_rank[7] += AD_col


            col = 23
            for i in range(8):
                write_in_xlsx(len(overall_rank[i]), col, overall_rank[i], ws_output)
                col += 1
            

            AE_col = ["","","Rank1 Octant ID"]
            for i in rank_list:
                x = i.index(min(i))
                AE_col.append(int_l[x])


            AF_col = ["","","Rank1 Octant Name"]
            for i in AE_col[3:]:
                AF_col.append(octant_name_id_mapping[str(i)])


            # Updating AF column ...
            AF_col_copy = list(AF_col[4:])
            write_in_xlsx(len(AF_col),32,AF_col,ws_output)


            final_mapping = {"Internal outward interaction":0, "External outward interaction":0, "External Ejection":0, "Internal Ejection":0, "External inward interaction":0, "Internal inward interaction":0, "Internal sweep":0, "External sweep":0}
            for item in AF_col_copy:
                if (item in final_mapping):
                    final_mapping[item] += 1
                else:
                    final_mapping[item] = 1


            final = []
            for i in AD_col[2:]:
                final.append(final_mapping[i])

        
            AE_col_ = ["","Count of Rank 1 Mod Values"] + final
            AE_col += AE_col_
            write_in_xlsx(len(AE_col),31,AE_col,ws_output)

            # ========================================== Tut05 Part-End ===============================================
            # ========================================== Tut02 Part-Start =============================================

            AH_col = ['','','']
            for i in range(n_ranges+1):
                AH_col.append("From")
                AH_col += ["" for i in range(13)]
            write_in_xlsx(len(AH_col), 34, AH_col, ws_output)


            AI_col = ["Overall Transition","","Octant #"] + st

            q = 0
            r = mod
            nr = n_ranges
            while nr > 0:   
                nr -= 1
                if r <= len(Oct[2:]):
                    if q==0:
                        p = f"{0000} - {r-1}"
                    else:
                        p = f"{q} - {r-1}"
                else:
                    p = f"{q} - {len(Oct[2:])}"
                AI_col += ["","","","Mod Transition Count"]
                AI_col.append(p)
                AI_col.append("Octant #")
                AI_col += st
                q = r 
                r = r + mod
            write_in_xlsx(len(AI_col), 35, AI_col, ws_output)


            # 8 new list are created to store the count value of {+1,-2,+2,-2,+3,-3,+4,-4}
            overall_count1 = []
            e = 0
            for i in range(8):
                empty_list = ["",""]
                empty_list.append(st[e])
                overall_count1.append(empty_list)
                e += 1


            for j in range(8):
                l = [0]*8
                d_f = df2
                for i in range(2,len(d_f)+1):  
                    if(d_f[i-1] == int_l[j]):
                        if(d_f[i-2] == 1):
                            l[0] += 1
                        elif(d_f[i-2] == -1):
                            l[1] += 1
                        elif(d_f[i-2] == 2):
                            l[2] += 1
                        elif(d_f[i-2] == -2):
                            l[3] += 1
                        elif(d_f[i-2] == 3):
                            l[4] += 1
                        elif(d_f[i-2] == -3):
                            l[5] += 1
                        elif(d_f[i-2] == 4):
                            l[6] += 1
                        elif(d_f[i-2] == -4):
                            l[7] += 1
                if j==0:
                    l+=["","","","","To"]
                else:
                    l += ["" for i in range(5)]

                overall_count1[j] += l  

            q = 0
            r = mod
            n_r_ = n_ranges
            while n_r_>0:
                n_r_ -= 1
                for j in range(8):
                    overall_count1[j].append(st[j])
                    l = [0]*8
                    d_f = df2[q:r+1]
                    for i in range(2,len(d_f)+1):  
                        if(d_f[i-1] == int_l[j]):
                            if(d_f[i-2] == 1):
                                l[0] += 1
                            elif(d_f[i-2] == -1):
                                l[1] += 1
                            elif(d_f[i-2] == 2):
                                l[2] += 1
                            elif(d_f[i-2] == -2):
                                l[3] += 1
                            elif(d_f[i-2] == 3):
                                l[4] += 1
                            elif(d_f[i-2] == -3):
                                l[5] += 1
                            elif(d_f[i-2] == 4):
                                l[6] += 1
                            elif(d_f[i-2] == -4):
                                l[7] += 1
                    if j==0 and (n_r_ != 0):
                        l+=["","","","","To"]
                    else:
                        l += ["" for i in range(5)]

                    overall_count1[j] += l  
                q = r 
                r += mod

            col = 36
            for i in range(8):
                write_in_xlsx(len(overall_count1[i]), col, overall_count1[i], ws_output)
                col += 1
            ws_output.cell(row=2, column=36).value = "To"


            # ========================================== Tut02 Part-End =================================================

            # ========================================== Tut03 Part-Start ===============================================


            g1 = ["Longest Subsquence Length"," ","Octant ##","+1","-1","+2","-2","+3","-3","+4","-4"]
            g2 = ["","","Longest Subsquence Length"]
            g2_ = []
            g3 = ["","","Count"]
            g3_ = []

            oct = Oct[2:]
            T = []
            for col in ws_output['A']:
                T.append(col.value)
            time1 = T[2:]
            n = len_H - 2

            # count1 to store initial count and previous_count is used to store previous count as the count value keeps on changing.
            for j in int_l:
                count1 = 0
                prev = 0
                indexend = 0
                for i in range(n):
                    if(oct[i] == j):
                        count1 += 1
                    else:
                        if(count1 > prev):
                            prev = count1
                            indexend = i
                        count1 = 0
                # c is used to store total number of times the small longest subsequence occurs
                c = 0
                count2 = 0
                for i in range(n):
                    if(oct[i] != j):
                        count2 = 0
                    else:
                        count2 += 1
                        if(count2 == prev):
                            c+=1
                            count2 = 0
                #print(prev, c)
                g2_.append(prev)
                g3_.append(c)
            
            #print(g2_,g3_)
            g2 = g2 + g2_
            g3 = g3 + g3_

            # 2d list to store entire small longest subsequence table
            g_final = []
            g_final.append(g1)
            g_final.append(g2)
            g_final.append(g3)

            col = 45
            for i in range(3):
                write_in_xlsx(len(g_final[i]), col, g_final[i], ws_output)
                col += 1

            # ========================================== Tut03 Part-End ===================================================
            
            # ========================================== Tut04 Part-Start =================================================

            tut_04(n,time1,oct,int_l,g3_,ws_output)


            # ========================================== Tut04 Part-End ===================================================

            # ========================================== Border - Start ==================================================



            set_border(ws_output, f'N3:AF{n_ranges+4}')
            set_border(ws_output, f'AC{n_ranges+6}:AE{n_ranges+14}')
            set_border(ws_output, 'AS3:AU11')

            column_start = 'AI'
            column_end = 'AQ'
            x = 3
            for i in range(n_ranges+1):
                start = column_start + f'{x}'
                end = column_end + f'{x+8}'
                set_border(ws_output, f'{start}:{end}')
                # print(start,end)
                x += 14
            
            y = sum(g3_) + 19
            set_border(ws_output, f'AW3:AY{y}')

            # ============================================ Border - End ==================================================

            # ============================================ Bgcolor - Start ===============================================

            bgcolor(f'W4:AD{n_ranges+4}','FFFF00',ws_output,0)

            column_start = 'AJ'
            column_end = 'AQ'
            x = 4
            for i in range(n_ranges+1):
                start = column_start + f'{x}'
                end = column_end + f'{x+7}'
                max_color(f'{start}:{end}', ws_output, 0)
                # print(start,end)
                x += 14

            # ============================================ Bgcolor - End =================================================

            current_time = datetime.now()

            if iter == 1:
                output_file_name = f.split(".xlsx")[0] + f"_{int(mod)}" + f"_{current_time.year}-{current_time.month}-{current_time.day}-{current_time.hour}-{current_time.minute}-{current_time.second}" + ".xlsx"
            else:
                output_file_name = filename.split(".xlsx")[0] + f"_{int(mod)}" + f"_{current_time.year}-{current_time.month}-{current_time.day}-{current_time.hour}-{current_time.minute}-{current_time.second}" + ".xlsx"
            
            try:
                wb_output.save(output_file_name)
            except:
                print("Error in generating output file")
    
except:
		print("Error in calling octant_analysis function")


try:
    def tut_04(n,time1,oct,int_l,g3_,ws_output):
        dk = []
        # g0,g1,g2,g3 are the list created which will store the values of the column 
        x = sum(g3_) + 18
    
        g1 = ["Longest Subsquence Length with Range","","Octant ###"]
        l_ = ["+1","-1","+2","-2","+3","-3","+4","-4"]
        for i in range(len(l_)):
            g1.append(l_[i])
            g1.append("Time")
            gap = [" " for _ in range(g3_[i])]
            g1 = g1 + gap

        g2 = ["","","Longest Subsquence Length"]
        g3 = ["","","Count"]
        k0 = []
        k1 = []
        l_0 = []
        l_1 = []

        for j in int_l:
            count1 = 0
            prev = 0
            for i in range(n):
                if(oct[i] == j):
                    count1 += 1
                else:
                    if(count1 > prev):
                        prev = count1
                        indexend = i
                    count1 = 0
            c = 0
            count2 = 0
            indexend = 0
            l0 = []
            l1 = []   
            for i in range(n):
                if(oct[i] != j):
                    count2 = 0
                else:
                    count2 += 1
                if(count2 == prev):
                    c+=1
                    count2 = 0
                    indexend = i
                    l0.append(indexend-prev+1)
                    l1.append(indexend)
         
            l_0.append(l0)
            l_1.append(l1)
            k0.append(prev)
            k1.append(c)

        for j in range(8):
            g2.append(k0[j])
            g2.append("From")
            z1 = l_0[j]
            for p in range(len(z1)):
                t = time1[z1[p]]
                g2.append(t)

        for j in range(8):
            g3.append(k1[j])
            g3.append("To")
            z2 = l_1[j]
            for p in range(len(z2)):
                t = time1[z2[p]]
                g3.append(t)
                
        dk.append(g1)
        dk.append(g2)
        dk.append(g3)

        col = 49
        for i in range(3):
            write_in_xlsx(len(dk[i]), col, dk[i], ws_output)
            col += 1
except:
	print("Error in calling tut_04 function")


global radio
try:
    def proj_octant_gui():
        global radio
        
        try:
            image = Image.open(r"C:\Users\Sahil\Desktop\Project\Code_IMG.jpg")
            st.image(image)
        except:
            print("Error in opening image.")

            
        col1, col2, col3 = st.columns([1,1,1])
        with col1:
            pass
        radio = col2.radio(label="Select Input Category", options=['Single xlsx file'])
        # radio = col2.radio(label="Select Input Category", options=['Single xlsx file', 'Bulk Conversion'])
        with col3:
            pass

        if radio == "Bulk Conversion":
            upload_path = col2.text_input("Enter location of input folder", "Paste Here ...")

            flag1 = False
            if len(upload_path) == 0:
                time.sleep(10)

            try:
                if upload_path:
                    upload_path = upload_path.replace('\\','/')
                    count = 0
                    for i in os.listdir(upload_path):
                        if re.search(".xlsx", i):
                            count +=1

                    if count == len(os.listdir(upload_path)):
                        flag1 = True
                    else:
                        st.warning('Input Address contains file other than xlsx files..."', icon="⚠️")
            except:
                print("Error occurred due to empty upload_path variable")

            if flag1:
                try:
                    os.chdir(upload_path)
                    files = os.listdir(upload_path)

                    data = files[0]

                    col4, col5 = st.columns(2)
                    if data[-5:] == ".xlsx":
                        col4.text(f"File Name... \"{data}\"")
                        df_display = pd.read_excel(data)
                        col4.dataframe(df_display)
                    if(len(files)>1):
                        col5.subheader("Other files in the directory ")
                        for i in files[1:]:
                            col5.text(f"File Name... \"{i}\"")

                    col_1, col_2, col_3 = st.columns(3)

                    mod = col_2.text_input(f"Enter Mod Value between (0, {df_display.shape[0]}]: ")
                    # Using regex to prevent unwanted input from user to cause any error ahead in the program.
                    if bool(re.search('[a-z]', mod, re.IGNORECASE)) or bool(re.search('[@_!#$%^&*()<>?/\|.}{~:.]', mod)):
                        st.warning('Enter Numeric Value Only..."', icon="⚠️")
                    else:
                        if len(list(map(int,mod.split()))) == 0:
                            pass
                        else:
                            mod = list(map(int,mod.split()))[0]
                            if mod < 0 or mod > df_display.shape[0]:
                                st.warning('Entered Mod value is either below than 0 or Greater than total number of rows present..."', icon="⚠️")
                    

                    col_1, col_2, col_3 = st.columns(3)
                    # Compute button
                    if(col_2.button('Compute')):
                        col_2.success(f"Mod = {mod}")

                        try:
                            octant_analysis(1,upload_path,int(mod))
                        except:
                            print("Error in calling octant_analysis function.")
                            

                    if col_2.checkbox("Display Output File"):
                        st.subheader("Select the Output File to be displayed below")
                        col6,col7 = st.columns(2)

                        current_dir = upload_path.rsplit("/",1)[0]
                        output_dir = current_dir + "/output3/"
                        os.chdir(output_dir)

                        output_files = os.listdir()
                        select_box = col7.radio(label="Which output file do you want to display", options=output_files)
                    
                        colors = display(select_box,col6,mod)

                        # Apply option for additiona features to be added to the main output file...
                        if st.sidebar.checkbox("Apply Additional Features"):
                            font_size = colors[-1]
                            del colors[-1]
                            additional_features(colors,font_size,mod)

                            zip_file(upload_path,1)
            

                except FileNotFoundError:
                    pass

        elif radio=="Single xlsx file":    
            uploaded_file = st.file_uploader("Upload Dataset",type="xlsx")

            try:
                if uploaded_file is not None:
                    # Screen is subdivided into three equal parts...
                    col_1, col_2, col_3 = st.columns(3)

                    if uploaded_file.type =="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
                        df=pd.read_excel(uploaded_file)
                        col_2.text(f"File Name... \"{uploaded_file.name}\"")
                        col_2.dataframe(df)


                    mod = col_2.text_input(f"Enter Mod Value between (0, {df.shape[0]}]: ")
                    # Using regex to prevent unwanted input from user to cause any error ahead in the program.
                    if bool(re.search('[a-z]', mod, re.IGNORECASE)) or bool(re.search('[@_!#$%^&*()<>?/\|.}{~:.]', mod)):
                        st.warning('Enter Numeric Value Only..."', icon="⚠️")
                    else:
                        if len(list(map(int,mod.split()))) == 0:
                            pass
                        else:
                            mod = list(map(int,mod.split()))[0]
                            if mod < 0 or mod > df.shape[0]:
                                st.warning('Entered Mod value is either below than 0 or Greater than total number of rows present..."', icon="⚠️")


                    col_1, col_2, col_3 = st.columns(3)
                    # Compute button
                    if(col_2.button('Compute')):
                        col_2.success(int(mod))
                        # Octant_analysis function called ...
                        octant_analysis(0,[df,uploaded_file.name],int(mod))


                    if col_2.checkbox("Display Output File"):

                        # Calling display function...
                        colors = display(os.listdir(os.getcwd().replace("\\","/"))[0],0,mod)

                        # Apply option for additiona features to be added to the main output file...
                        # if st.sidebar.checkbox("Apply Additional Features"):
                        #     font_size = colors[-1]
                        #     del colors[-1]
         
                        #     additional_features(colors,font_size,mod)                        
    
                        #     # Zipping files...
                        #     if os.getcwd().rsplit("\\")[-1] == "output3":
                        #         current_dir = os.getcwd()[:os.getcwd().index('\output3')].replace("\\","/")
                        #     else:
                        #         current_dir = os.getcwd().replace("\\","/")


                        #     zip_file(current_dir,0)
                            

            except TypeError:
                pass
except:
    print("Error in calling proj_octant_gui")

# additional_features function is called to apply background color and color to the max values
try:
    def additional_features(colors,font_size,mod):
        global radio

        if radio == "Bulk Conversion":
            list_of_files = os.listdir()
        else:
            list_of_files = os.listdir(os.getcwd().replace("\\","/"))

        for c in range(len(colors)):
            colors[c] = colors[c][1:]


        for file in list_of_files:
            wb = openpyxl.load_workbook(file)
            ws = wb.active


            # Calculation number of ranges required.
            n = len(ws['A'])-2
            # if n%mod == 0:
            n_ranges = int(n//mod)
            # else:
            #     n_ranges = math.ceil(n/mod)

            
            bgcolor('N3:AF3',colors[0],ws,1)
            bgcolor(f'N4:N{n_ranges+4}',colors[0],ws,2)

            bgcolor(f'AC{n_ranges+6}:AE{n_ranges+6}',colors[0],ws,1)
            bgcolor('AC11:AC18',colors[0],ws,2)

            x = 3
            for j in range(n_ranges+1):
                bgcolor(f'AI{x}:AQ{x}',colors[0],ws,1)
                bgcolor(f'AI{x+1}:AQ{x+8}',colors[0],ws,2)
                x += 14

            bgcolor(f'AS3:AU3',colors[0],ws,1)
            bgcolor(f'AS4:AS11',colors[0],ws,2)

            y=19
            for row in ws['AU4:AU11']:
                y += row[0].value
            
            bgcolor(f'AW3:AY3',colors[0],ws,1)
            bgcolor(f'AW4:AW{y}',colors[0],ws,2)


            bgcolor(f'O4:AF{n_ranges+4}',colors[1],ws,3)

            bgcolor(f'AD{n_ranges+7}:AE{n_ranges+14}',colors[1],ws,3)
            bgcolor('AT4:AU11',colors[1],ws,3)

            column_start = 'AJ'
            column_end = 'AQ'
            x = 4
            for i in range(n_ranges+1):
                start = column_start + f'{x}'
                end = column_end + f'{x+7}'
                bgcolor(f'{start}:{end}',colors[1],ws,3)
                x += 14
            
            bgcolor(f'AX4:AY{y}',colors[1],ws,3)

            bgcolor(f'W4:AD{n_ranges+4}',colors[2],ws,0)
            max_color(f'O4:V{n_ranges+4}',ws,colors[4])

            column_start = 'AJ'
            column_end = 'AQ'
            x = 4
            for i in range(n_ranges+1):
                start = column_start + f'{x}'
                end = column_end + f'{x+7}'
                max_color(f'{start}:{end}', ws, colors[3])
                # print(start,end)
                x += 14

            fontsize(font_size,ws)

            wb.save(file)
            print("Completed")
except:
	print("Error in calling additional_features function")
        


try:
    def display(option,col7,mod):
        pd.set_option('display.max_colwidth', 0)
        # print(os.getcwd())
        df = pd.read_excel(option,header=None)
        df.fillna('', inplace=True)

        if col7 != 0:
            col7.dataframe(df)
        else:
            st.dataframe(df)


        st.subheader("Octant Ranking Table")

        n = df.shape[0]

        # if n%mod == 0:
        n_ranges = int(n//mod)
        #     print(mod, 1)
        # else:
        #     n_ranges = math.ceil(n/mod)

        # Separating the Octant rank table from dataframe and storing it in oct_rank variable
        oct_rank = {}
        col = 13
        k=1
        
        for i in df.iloc[2,13:32].to_list():
            if i not in oct_rank:
                if k>=2 and k<=9:
                    i = str(int(float(i)))
                    oct_rank[i] = []
                    k+=1
                else:
                    oct_rank[i] = []

            if col==13 or col == 31:
                k += 1
                for j in range(n_ranges+1):
                    oct_rank[i].append(df.iloc[j+3,col])
            else:  
                for j in range(n_ranges+1):  
                    oct_rank[i].append(int(float(df.iloc[j+3,col])))
            
            col += 1

        df1 = pd.DataFrame(oct_rank)

        list_of_max_index = []
        for i in range(n_ranges+1):
            max1 = list(df1.iloc[i,1:9]).index(max(df1.iloc[i,1:9]))
            dup = [i,max1]
            list_of_max_index.append(dup)

        st.sidebar.text("Additional Features...")


        # Widgets...
        header = st.sidebar.color_picker('Pick a Color for Header of the Table', '#FFFFFF')
        body = st.sidebar.color_picker('Pick a Color for body of the Table', '#FFFFFF')
        rank1 = st.sidebar.color_picker('Pick a Color for Highlighing Rank1', '#FFFF00')
        max_color = st.sidebar.color_picker('Pick a Color for Max Values of Transition Count Tables','#FFFF00')
        max_color1 = st.sidebar.color_picker('Pick a Color for Max Values of Rank Tables','#FFFFFF')
        if body != "#FFFFFF" and max_color1 == "#FFFFFF":
            max_color1 = body
        font_size = st.sidebar.slider('Set the Font Size',value=10,min_value=10,max_value=20,step = 1)
        # ===========================================================================================

        # This part is used to display Octant rank table =============================================

        color_ = []
        for i in range(n_ranges+1):
            dup = []
            for j in range(len(df1.iloc[i,1:9])):
                flag = False
                for k in list_of_max_index:
    
                    if i == k[0] and j == k[1]:
                        dup.append(max_color1)
                        flag = True
                        break

                if flag == False:
                    dup.append(body)

            color_.append(dup)

        color_ = np.transpose(color_)

    
        list_of_min_index = []
        for i in range(n_ranges+1):
            min1 = list(df1.iloc[i,9:17]).index(min(df1.iloc[i,9:17]))
            dup = [i,min1]
            list_of_min_index.append(dup)

        color_min = []

        # Applying color to rank1 octants
        for i in range(n_ranges+1):
            dup = []
            for j in range(len(df1.iloc[i,9:17])):
                flag = False
                for k in list_of_min_index:
                    # print(i,j,k)
                    if i == k[0] and j == k[1]:
                        dup.append(rank1)
                        flag = True
                        break
                    # else:
                if flag == False:
                    dup.append(body)

            color_min.append(dup)

        color_min = np.transpose(color_min)


        go_fig1 = go.Figure()
        fill_color = []
        map_color = {1:f"{rank1}", 2:f"{body}", 3:f"{body}", 4:f"{body}", 5:f"{body}", 6:f"{body}", 7:f"{body}", 8:f"{body}"}
        
        iter1 = 0 # to add color_ list in fill_Color
        iter2 = 0
        for i in ['Octant ID', '1', '-1', '2', '-2', '3', '-3', '4', '-4','Rank Octant 1', 'Rank Octant -1', 'Rank Octant 2', 'Rank Octant -2','Rank Octant 3', 'Rank Octant -3', 'Rank Octant 4', 'Rank Octant -4', 'Rank1 Octant ID', 'Rank1 Octant Name']:
            if i in ['1', '-1', '2', '-2', '3', '-3', '4', '-4']:
                fill_color.append(color_[iter1])
                iter1 += 1

            elif i in ['Rank1 Octant ID', 'Rank1 Octant Name']:
                fill_color.append([f'{body}']*len(df1))
            elif i=='Octant ID':
                fill_color.append([f'{header}']*len(df1))
            else:
                fill_color.append(color_min[iter2])
                iter2 += 1


        # Plotly tables are used to display octant tables...

        obj = go.Table(columnwidth = [1000,800],header = dict(values=list(df1[['Octant ID', '1', '-1', '2', '-2', '3', '-3', '4', '-4','Rank Octant 1', 'Rank Octant -1', 'Rank Octant 2', 'Rank Octant -2','Rank Octant 3', 'Rank Octant -3', 'Rank Octant 4', 'Rank Octant -4', 'Rank1 Octant ID', 'Rank1 Octant Name']].columns),fill_color=header,
                    align='center',font = dict(color = '#051c2c'))
        ,cells = dict(values=[df1['Octant ID'], df1['1'], df1['-1'], df1['2'], df1['-2'], df1['3'], df1['-3'], df1['4'], df1['-4'],df1['Rank Octant 1'], df1['Rank Octant -1'], df1['Rank Octant 2'], df1['Rank Octant -2'],df1['Rank Octant 3'], df1['Rank Octant -3'], df1['Rank Octant 4'], df1['Rank Octant -4'], df1['Rank1 Octant ID'], df1['Rank1 Octant Name']],fill_color= fill_color,
                align='center',font = dict(color = '#051c2c')))

        go_fig1.add_trace(obj)
        go_fig1.update_layout(title_text=f"Octant Ranking Table for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5, margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size), width = 850)
        st.write(go_fig1)

        # =====================================================
        # =====================================================


        oct_rank1 = {'Octant ID':df.iloc[10:18,28],'Octant Name':df.iloc[10:18,29],'Count of Rank 1 Mod Values':df.iloc[10:18,30]}
        df2 = pd.DataFrame(oct_rank1)
        fill_color1 = []
        for i in ['Octant ID', 'Octant Name', 'Count of Rank 1 Mod Values']:
            if i=='Octant ID':
                fill_color1.append([f'{header}']*len(df2))
            else:
                fill_color1.append([f'{body}']*len(df2))
        
        go_fig2 = go.Figure()
        obj = go.Table(columnwidth = [100,800],header = dict(values=list(df2[['Octant ID', 'Octant Name', 'Count of Rank 1 Mod Values']].columns),fill_color=header,
                    align='center',font = dict(color = '#051c2c'))
        ,cells = dict(values=[df2['Octant ID'], df2['Octant Name'], df2['Count of Rank 1 Mod Values']],fill_color= fill_color1,
                align='center',font = dict(color = '#051c2c')))

        go_fig2.add_trace(obj)
        go_fig2.update_layout(margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
        st.write(go_fig2)


        # ==================================================================================================================

        # ============================================= Transition Count Table =============================================

        st.subheader("Transition Count Table")


        x = 3
        for j in range(n_ranges+1):
            col = ['Octant #','+1','-1','+2','-2','+3','-3','+4','-4']

            # Again making dictionary and then converted into dataframe which makes it much much easier to build a table using plotly.

            oct_rank1 = {'Octant #':[int(float(i)) for i in df.iloc[x:x+8,34]],'+1':[int(float(i)) for i in df.iloc[x:x+8,35]],'-1':[int(float(i)) for i in df.iloc[x:x+8,36]],'+2':[int(float(i)) for i in df.iloc[x:x+8,37]],'-2':[int(float(i)) for i in df.iloc[x:x+8,38]]
            ,'+3':[int(float(i)) for i in df.iloc[x:x+8,39]],'-3':[int(float(i)) for i in df.iloc[x:x+8,40]],'+4':[int(float(i)) for i in df.iloc[x:x+8,41]],'-4':[int(float(i)) for i in df.iloc[x:x+8,42]]}

            df3 = pd.DataFrame(oct_rank1)
            x += 14
            p = 0
            q = mod

            # list_of_max1 = []
            # for i in range(8):
            #     list_of_max1.append(max(df3.iloc[i,1:9]))

            list_of_max_index = []
            for i in range(8):
                max1 = list(df3.iloc[i,1:9]).index(max(df3.iloc[i,1:9]))
                dup = [i,max1]
                list_of_max_index.append(dup)

            # print(list_of_max_index)
            fill_color1 = []

            #  Getting position of max values in the dataframe and applying max color provided by user.
            color_ = []
            for i in range(8):
                dup = []
                for d in range(len(df3.iloc[i,1:9])):
                    flag = False
                    for k in list_of_max_index:
                       
                        if i == k[0] and d == k[1]:
                            dup.append(max_color)
                            flag = True
                            break
                        
                    if flag == False:
                        dup.append(body)

                color_.append(dup)
            color_ = np.transpose(color_)


            iter = 0 # to add color_ list in fill_Color
            for i in col:
                if i=='Octant #':
                    fill_color1.append([f'{header}']*len(df3))
                else:
                    fill_color1.append(color_[iter])
                    iter += 1


            go_fig3 = go.Figure()
            obj = go.Table(columnwidth = [100,200],header = dict(values=list(df3[['Octant #','+1','-1','+2','-2','+3','-3','+4','-4']].columns),fill_color=header,
                        align='center',font = dict(color = '#051c2c'))
            ,cells = dict(values=[df3['Octant #'],df3['+1'],df3['-1'],df3['+2'],df3['-2'],df3['+3'],df3['-3'],df3['+4'],df3['-4']],fill_color= fill_color1,
                    align='center',font = dict(color = '#051c2c')))

            go_fig3.add_trace(obj)
            if j == 0:
                go_fig3.update_layout(title_text=f"Overall Transition Count Table for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5,margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
            elif j==1:
                go_fig3.update_layout(title_text=f"Mod Transition Count Table from {'0'}-{q-1} for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5,margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
            elif j==n_ranges:
                go_fig3.update_layout(title_text=f"Mod Transition Count Table from {p}-{df.shape[0]} for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5,margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
            elif j>1 and j<n_ranges:
                go_fig3.update_layout(title_text=f"Mod Transition Count Table from {p}-{q-1} for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5,margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
            st.write(go_fig3)

            col8,col9,col10 = st.columns(3)
            for i in range(1,len(col)):
                if i<=4:
                    col8.text(f"Sum of \"{col[i]}\"... {sum(oct_rank1[col[i]])}")
                else:
                    col9.text(f"Sum of \"{col[i]}\"... {sum(oct_rank1[col[i]])}")
            col10.text(f"Total sum of the table is {sum(oct_rank1[col[1]])+sum(oct_rank1[col[2]])+sum(oct_rank1[col[3]])+sum(oct_rank1[col[4]])+sum(oct_rank1[col[5]])+sum(oct_rank1[col[6]])+sum(oct_rank1[col[7]])+sum(oct_rank1[col[8]])}")    

            p = q
            q += mod

        # ===================================================================================================================

        # ============================================= Longest Subsequence Table ===========================================

        st.subheader("Longest Subsquence Length Table")

        long_len = {"Octant ##":df.iloc[3:11,44],"Longest Subsquence Length":df.iloc[3:11,45],"Count":[int(float(i)) for i in df.iloc[3:11,46]]}
        df4 = pd.DataFrame(long_len)
        # print(long_len)

        fill_color3 = []
        for i in long_len.keys():
            if i=='Octant ##':
                fill_color3.append([f'{header}']*len(df4))
            else:
                fill_color3.append([f'{body}']*len(df4))

        go_fig4 = go.Figure()
        obj = go.Table(columnwidth = [100,800],header = dict(values=list(df4[['Octant ##', 'Longest Subsquence Length', 'Count']].columns),fill_color=header,
                    align='center',font = dict(color = '#051c2c'))
        ,cells = dict(values=[df4['Octant ##'], df4['Longest Subsquence Length'], df4['Count']],fill_color= fill_color3,
                align='center',font = dict(color = '#051c2c')))

        go_fig4.add_trace(obj)
        go_fig4.update_layout(title_text=f"Longest Subsquence Length Table for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5, margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
        st.write(go_fig4)

        # ===============================================================================================================
        
        # ===================================Longest Subsequence Table With Raange ======================================


        st.subheader("Longest Subsquence Length with Range Table")

        long_len_r = {"Octant ###":df.iloc[3:sum(long_len['Count'])+19,48],"Longest Subsquence Length":df.iloc[3:sum(long_len['Count'])+19,49],"Count":[int(float(i)) if i!="To" else i for i in df.iloc[3:sum(long_len['Count'])+19,50]]}
        df5 = pd.DataFrame(long_len_r)
        # print(long_len)

        fill_color3 = []
        for i in long_len_r.keys():
            if i=='Octant ###':
                fill_color3.append([f'{header}']*len(df5))
            else:
                fill_color3.append([f'{body}']*len(df5))

        go_fig4 = go.Figure()
        obj = go.Table(columnwidth = [100,500],header = dict(values=list(df5[['Octant ###', 'Longest Subsquence Length', 'Count']].columns),fill_color=header,
                    align='center',font = dict(color = '#051c2c'))
        ,cells = dict(values=[df5['Octant ###'], df5['Longest Subsquence Length'], df5['Count']],fill_color= fill_color3,
                align='center',font = dict(color = '#051c2c')))

        go_fig4.add_trace(obj)
        go_fig4.update_layout(title_text=f"Longest Subsquence Length with Range Table for... \"{option.rsplit('.',1)[0]}\" ", title_x=0.5, margin = dict(t=50, l=25, r=25, b=25),font = dict(color = header,size = font_size))
        st.write(go_fig4)


        # returning the selected colors and font_Size
        return [header,body,rank1,max_color,max_color1,font_size]

except:
    print("Error in calling display function.")




from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")



proj_octant_gui()


#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
